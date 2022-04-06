import requests
from fake_useragent import UserAgent
import time
from openpyxl import Workbook

class TencentVirus:
    def __init__(self):
        self.url = "https://api.inews.qq.com/newsqa/v1/automation/foreign/country/ranklist"
        self.headers = {
            'User-Agent': UserAgent().random
        }
        self.time = time.strftime("%Y-%m-%d")
        self.wb = Workbook()

    def get(self):
        try:
            html = requests.get(url=self.url, headers=self.headers).json()
        except:
            pass
        self.parse(html)


    def parse(self, data):
        data = data['data']
        sheet = self.wb.create_sheet('tenct')
        c = 0
        item = ["国家名称", "区域", "新增病例", "累计确诊", "死亡人数", "治愈人数", "现有确诊"]
        sheet.append(item)
        for i in range(len(data)):
            if c < len(data):
                c += 1
                sheet.cell(row=c+1, column=1, value=data[i]['name'])
                sheet.cell(row=c+1, column=2, value=data[i]['continent'])
                sheet.cell(row=c+1, column=3, value=data[i]['confirmAdd'])
                sheet.cell(row=c+1, column=4, value=data[i]['confirm'])
                sheet.cell(row=c+1, column=5, value=data[i]['dead'])
                sheet.cell(row=c+1, column=6, value=data[i]['heal'])
                sheet.cell(row=c+1, column=7, value=data[i]['nowConfirm'])
        self.wb.save('{}.xlsx'.format(self.time))


if __name__ == '__main__':
    tenct = TencentVirus()
    tenct.get()